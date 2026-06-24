#!/usr/bin/env python3
"""
sync_l10n.py  —  Import translations from the local xlsx into locales/*.json

Usage:
    python sync_l10n.py

Workflow:
    1. Writer / translators edit  "Subwoofer Localization.xlsx"
    2. Drop their updated copy into this folder, overwriting the existing file
    3. Run this script — it diffs and updates locales/*.json automatically

Requirements:
    pip install openpyxl
"""

import json, os, sys

HERE       = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(HERE, 'Subwoofer_Localization.xlsx')
LOCALE_DIR = os.path.join(HERE, 'locales')

# Column indices in the sheet (0-based, matching the xlsx structure)
# A=key  B=category  C=context  D=source  E=en
LANGS    = ['en', 'zh-CN']
COL_KEY  = 0
COL_EN   = 4   # English is column E (index 4)
# zh-CN is column G (index 6), ja=F(5), zh-CN=G(6)

def load_locale(lang):
    path = os.path.join(LOCALE_DIR, f'{lang}.json')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_locale(lang, data):
    path = os.path.join(LOCALE_DIR, f'{lang}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')

def is_real_key(key):
    """Locale keys use dot notation. Category headers are ALL CAPS with no dots."""
    if not isinstance(key, str) or '.' not in key:
        return False
    stripped = key.strip()
    # Category group headers are ALL CAPS (e.g. "UI CHROME") — skip them
    return stripped != stripped.upper()

def sync():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('Run:  pip install openpyxl')
        sys.exit(1)

    if not os.path.exists(XLSX_PATH):
        print(f'ERROR: spreadsheet not found at:\n  {XLSX_PATH}')
        sys.exit(1)

    print(f'Reading {os.path.basename(XLSX_PATH)} …')
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['All Strings']

    existing = {lang: load_locale(lang) for lang in LANGS}
    updated  = {lang: dict(existing[lang]) for lang in LANGS}
    changes  = {lang: {} for lang in LANGS}
    added    = {lang: {} for lang in LANGS}

    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:          # skip header
            first_row = False
            continue

        if not row:
            continue           # skip truly empty rows
        key = row[COL_KEY] if len(row) > COL_KEY else None
        if not is_real_key(key):
            continue           # skip category group headers and blank rows

        for lang_idx, lang in enumerate(LANGS):
            col   = COL_EN + lang_idx
            value = str(row[col]).strip() if row[col] is not None else ''

            if not value:
                continue       # empty cell → leave existing value alone

            old = existing[lang].get(key, '')

            if value != old:
                updated[lang][key] = value
                if old:
                    changes[lang][key] = (old, value)
                else:
                    added[lang][key] = value
            else:
                updated[lang][key] = value

    wb.close()

    # Write
    for lang in LANGS:
        save_locale(lang, updated[lang])

    # Report
    total = sum(len(changes[l]) + len(added[l]) for l in LANGS)
    if total == 0:
        print('✓ No changes — locale files are already up to date.')
        return

    print()
    for lang in LANGS:
        n_changed = len(changes[lang])
        n_added   = len(added[lang])
        if not n_changed and not n_added:
            continue
        parts = []
        if n_changed: parts.append(f'{n_changed} updated')
        if n_added:   parts.append(f'{n_added} new')
        print(f'  {lang:8s}  {", ".join(parts)}')
        for key, (old, new) in list(changes[lang].items())[:3]:
            print(f'           {key}')
            print(f'             was: {(old[:70] + "…") if len(old) > 70 else old}')
            print(f'             now: {(new[:70] + "…") if len(new) > 70 else new}')
        if n_changed > 3:
            print(f'           … and {n_changed - 3} more')

    print()
    print(f'✓ {total} change(s) written to locales/')

if __name__ == '__main__':
    sync()
